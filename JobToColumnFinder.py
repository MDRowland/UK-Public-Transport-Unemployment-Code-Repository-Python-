import openpyxl as opxl

wb = opxl.load_workbook("ExcelSheets/SCR_MSOA_Workplace_Data_Split.xlsx")
ws=wb['SCR_Workplace_Data']
k=5
f=open("Programme Files/TargetJobCodes.txt")
JobCodes = eval(f.readline())
f.close()
Rows=[]
while k<ws.max_column:
    if (int(list(ws.cell(row=1,column=k).value.split())[1])) in JobCodes:
        Rows.append(ws.cell(row=1,column=k).column)
    k+=1
f=open("Programme Files/TargetJobColumns.txt","w")
f.write(str(Rows))
f.close()
