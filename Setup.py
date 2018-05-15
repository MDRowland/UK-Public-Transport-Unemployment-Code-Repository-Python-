import openpyxl as opxl
import os
####################################
####### Generic Functions ##########
def file_len(fname):              ##
    with open(fname) as f:        ##
        for i, l in enumerate(f): ##
            pass                  ##
    return i + 1                  ##
####################################
####### Files and other ############

print('Performing first-time setup.')
if os.path.exists("Output Files"):
    pass
else:
    print("Generating Directory Structure")
    os.makedirs("Output Files")

if os.path.exists("Programme Files"):
    pass
else:
    os.makedirs("Programme Files")
    f = open("Programme Files/z.txt", "w")
    f.write("0 0")
    f.close()
    f = open("Programme Files/CostCoefficients.txt", "w")
    f.write("{'Time': [0, 1, 0, 0, 0, 0], 'Disabled': [0.1, 1, 0, 0.3, 0.1, 0.3], 'Time + Living Wage': [459.77, 1, 0, 0, 0, 0], 'Cycling': [0, 0, 1, 0, 0, 0]}")
    f.close()
    f = open("Programme Files/CostCoefficients_MetaData.txt", "w")
    f.write("Name Cost_Coefficent Time_Coefficent Distance_Coefficent Changes_malus walking_time_malus walking_distance_malus")
    f.close()
    f = open("Programme Files/MaxCost.txt", "w")
    f.write("{'Twenty Minutes': 1200, 'One Hour': 3600, 'Ten Miles': '16000'}")
    f.close()
    f = open("Programme Files/TargetJobCodes.txt", "w")
    f.write("[927,622,613,923,711,624,924,543,913,911]")
    f.close()
if os.path.exists("Locality Files"):
    pass
else:
    os.makedirs("Locality Files")   
if os.path.exists("ExcelSheets"):
    pass
else:
    os.makedirs("ExcelSheets")
    for sheet in ["SCR_GeoCode_DeprivationFraction","SCR_LSOA_Estimates","SCR_LSOA_Workplace_Totals","SCR_LSOAConversions","SCR_MSOA_Workplace_Data_Split"]:
        os.rename(sheet,str("ExcelSheets/"+sheet))
    
conversion_file = "ExcelSheets/SCR_LSOA_Estimates.xlsx"
wb = opxl.load_workbook(conversion_file, read_only = True)

print("Identifying numbers of LSOAs and MSOAs.")
LSOA_dictionary = {}
MSOA_dictionary = {}
places =wb.get_sheet_names()
for place in places:
    LSOA_dictionary[place]=(wb[place].max_row-5)
    n=1
    temp=[]
    while n<(wb[place].max_row-5):
        if wb[place]["C"+str(n+5)].value[:-1] in temp:
            pass
        else:
            temp.append(wb[place]["C"+str(n+5)].value[:-1])
        n+=1
    MSOA_dictionary[place]=int(len(temp))
    print(place,"complete")
f =open("Programme Files/LSOA_Dictionary.txt", "w")
f.write(str(LSOA_dictionary))
f.close()
f =open("Programme Files/MSOA_Dictionary.txt", "w")
f.write(str(MSOA_dictionary))
f.close()
print('LSOAs counted.')
print('Setup complete!')
