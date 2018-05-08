import openpyxl as opxl
import os
####################################
####### Generic Functions ##########
def file_len(fname):              ##
    i=0           ##
    with open(fname) as f:        ##
        for i, l in enumerate(f): ##
            pass                  ##
    return i + 1                  ##
####################################
####### Files and other ############
data_source = "ExcelSheets/SCR_GeoCode_DeprivationFraction.xlsx"
pop_data_source = "ExcelSheets/SCR_LSOA_Estimates.xlsx"
MSOA_job_data_source = "ExcelSheets/SCR_MSOA_Workplace_Data_Split.xlsx"
LSOA_job_data_source = "ExcelSheets/SCR_LSOA_Workplace_Totals.xlsx"
translation_data_source = "ExcelSheets/SCR_LSOAConversions.xlsx"
thresholds= [1,5,10,20,25,30,40,50,60,70,75,80,90,95,99,100]
####################################

wb = opxl.load_workbook(data_source, read_only=True)
wb2 = opxl.load_workbook(pop_data_source, read_only=True)
wb3 = opxl.load_workbook(MSOA_job_data_source, read_only=True)
wb4 = opxl.load_workbook(translation_data_source, read_only=True)
wb5= opxl.load_workbook(LSOA_job_data_source, read_only=True)    #Identifying loctions, setting constants etc etc
ws=wb[placename]
ws4=wb4[placename]
Mid2016Persons = wb2[placename]

lengthPop = Mid2016Persons.max_row
lengthRRLLData = ws.max_row
lengthConvert = ws4.max_row
f=open("Programme Files/TargetJobColumns.txt")
JobColumns=eval(f.readline())
f.close()
n=1
#Testing for existing data and creating paths where necessary 
if os.path.exists("Locality Files/"+str(placename)+"_Nodes"):
    if os.path.exists("Locality Files/"+str(placename)+"_Nodes\LSOAs.txt"):
        if os.path.exists("Locality Files/"+str(placename)+"_Nodes\MSOAs.txt") and nodetype == 'MSOA':
            print(str(placename)+" LSOA and MSOA data found. Checking validity...")
            NeedLSOAs = False
            NeedMSOAs = False
        elif nodetype != 'MSOA':
            NeedMSOAs = False
            NeedLSOAs = False
        else:
            print(str(placename)+" LSOA data does exist, but MSOA does not. Creating MSOA data...")
            NeedMSOAs = True
            NeedLSOAs = False
    else:
        print("No data for "+str(placename)+" on file- creating now.")
        NeedLSOAs = True
        if nodetype == 'MSOA':
            NeedMSOAs = True
else:
    os.makedirs("Locality Files/"+str(placename)+"_Nodes")
    NeedLSOAs = True
    if nodetype == 'MSOA':
        NeedMSOAs = True
    else:
        NeedMSOAs = False
numberLSOA, numberMSOA = LSOA_dictionary[placename],MSOA_dictionary[placename] #These dictionaries, created by Setup.py have the numbers of L&MSOAs per area 
print("Checking "+str(placename)+" LSOAs...")
print("There are "+str(numberLSOA)+" LSOAs in "+str(placename))
LSOA_Array = []
if NeedLSOAs == False and file_len("Locality Files/"+str(placename)+"_Nodes/LSOAs.txt") ==numberLSOA: # If LSAOs are correct and unneeded, then they are retrieved from file.
    print("LSOA data passes consistency check. Retrieving.")                                          # "Correct" means does file length match figure retrieved from the dictionaries above
    LSOA_Output_File = open("Locality Files/"+str(placename)+"_Nodes/LSOAs.txt", "r")
    k = 0
    lines = LSOA_Output_File.readlines()
    while k<numberLSOA:
        cl = lines[k].split()
        LSOA_Array.append(cl)
        k+=1
else:           # Otherwise, they are found and the data is created.
    print("LSOA data incorrect or outdated. Updating...")
    LSOA_Output_File = open("Locality Files/"+str(placename)+"_Nodes/LSOAs.txt", "w")
    n=1
    print("Finding LSOA data...")
    while n < lengthRRLLData:      
        current_node = (ws[str("A"+str(n+1))].value)
        MSOA_name = (ws[str("B"+str(n+1))].value)[:-1]
        currlong = (ws[str("E"+str(n+1))].value)
        currlat = (ws[str("F"+str(n+1))].value)
        k = 1
        while k<lengthPop:
            if str(Mid2016Persons[str('A'+str(k))].value)==str(current_node):
                Popline = Mid2016Persons[str('A'+str(k))].row           #Finds line in LSOA estimates file corresponding to current node
                break
            k+=1
        k=1   
        while k<wb5[placename].max_row:
            if str(wb5[placename][str('C'+str(k))].value)==str(current_node):
                LSOA_total = wb5[placename]['D'+str(k)].value
                break
            k+=1
        k=1
        
        while k<wb3[placename].max_row:
            if str(wb3[placename][str('B'+str(k))].value)==str(MSOA_name):
                MSOA_total_jobs = wb3[placename]['D'+str(k)].value
                target_jobs=0
                for col in JobColumns:
                    target_jobs+=wb3[placename][col+str(k)].value
                break
            k+=1
        projected_jobs=round((LSOA_total*target_jobs)/MSOA_total_jobs)
        
        
        if round((n/numberLSOA)*100)in thresholds:
            if round((n/lengthPop)*100) != round(((n-1)/lengthPop)*100):
                print("    "+str("{:5.2f}".format(round((n/lengthPop)*100,2)))+"% Complete") # Prints out progress percentage for benifit of user  
        LSOA_Output_File.write(str(current_node)+' '+str(Mid2016Persons[str('D'+str(Popline))].value)+' '+str(ws['D'+str(n+1)].value)+' '+str(currlong)+' '+str(currlat)+' '+str(projected_jobs)+"\n")
        # Adds the current LSOA data to the Array
        LSOA_Array.append([str(current_node),str(Mid2016Persons[str('D'+str(Popline))].value),str(ws['E'+str(n+1)].value),str(currlong),str(currlat),str(projected_jobs)])
        n+=1
if os.path.exists("Locality Files/"+str(placename)+"_Nodes/NodeMetaData.txt")==False:
    print("    Metadata does not exist- creating.")
    metadata = open("Locality Files/"+str(placename)+"_Nodes/NodeMetaData.txt", "w")
    metadata.write("Node_(LSOA)_Code Total_Population Employment_Rate_(Unemployment_fraction) LSOA_Longitude LSOA_Latitude Target_Job_Number")
    metadata.close()
    print("    Metadata created.") #Metadata Creation
LSOA_Output_File.close()
######################################################## MSOAs ################################################################
print("There are "+str(numberMSOA)+" MSOAs in "+str(placename))
if NeedMSOAs==False:
        if nodetype == 'MSOA':
            
            #If MSOAs are being used as the nodes in the modeling, this will find the MSOAs. Else See above.
            print("Checking MSOA data...")
            n=1
            k=0
            if numberMSOA==file_len("Locality Files/"+str(placename)+"_Nodes\MSOAs.txt"):
                print('MSOAs are on file and correct.') #retrive?
            else:
                print('MSOAs are incorrect or outdated. Updating.')
                NeedMSOAs = True            
if NeedMSOAs==True:
        n=0
        MSOA_Array = []
        print("Finding MSOAs...")
        while n<numberLSOA:
            
            CurrLSOA=LSOA_Array[n][0]
            k=1
            while k<lengthConvert:   # Goes down all possible LSOAs and retireives corresponding MSOA
                if ws4[str("B"+str(k))].value == CurrLSOA:
                    CurrMSOA = ws4[str("D"+str(k))].value
                    break
                k+=1
            
            for x in range(len(MSOA_Array)):
                #print(x)#Tests Current MSOA under consideration to see if has already been found
                if CurrMSOA in MSOA_Array[x]:
                  #If it is, adds population and unemployment fraction
                    MSOA_Array[x][2]=(((int(LSOA_Array[n][1])*float(LSOA_Array[n][2]))+(int(MSOA_Array[x][1])*float(MSOA_Array[x][2])))/(int(MSOA_Array[x][1])+int(LSOA_Array[n][1])))
                    MSOA_Array[x][1]=int(LSOA_Array[n][1])+int(MSOA_Array[x][1])
                    MSOA_Array[x][3]=float(MSOA_Array[x][3])+(float(LSOA_Array[n][3])-float(MSOA_Array[x][3]))*(int(LSOA_Array[n][1])/int(MSOA_Array[x][1]))
                    MSOA_Array[x][4]=float(MSOA_Array[x][4])+(float(LSOA_Array[n][4])-float(MSOA_Array[x][4]))*(int(LSOA_Array[n][1])/int(MSOA_Array[x][1]))
                  #The latter two move the current population weighted centroid according to the population in the MSOA so far and the LSOA being added
                    break

            else:   #for-else structure
                    #If this is a first time, the MSOA is just the LSOA values under the MSOA code
                    k=1
                    while k<wb3[placename].max_row:
                        if str(wb3[placename][str('C'+str(k))].value)==CurrMSOA:
                            target_jobs=0
                            for col in JobColumns:
                                target_jobs+=wb3[placename][col+str(k)].value #Finds "target" jobs 
                            break
                        k+=1
                    MSOA_Array.append([str(CurrMSOA),LSOA_Array[n][1],LSOA_Array[n][2],LSOA_Array[n][3],LSOA_Array[n][4],target_jobs])
            if round((n/numberLSOA)*100)in thresholds:
                if round((n/numberLSOA)*100) != round(((n-1)/numberLSOA)*100):
                    print("    "+str("{:5.2f}".format(round((n/numberLSOA)*100,2)))+"% Complete") # Prints out progress percentage for benifit of user
            
            n+=1
        x=0
        with open("Locality Files/"+str(placename)+"_Nodes\MSOAs.txt", "w") as MSOA_Output_File:
            while x<len(MSOA_Array):
                MSOA_Output_File.write(str(MSOA_Array[x][0])+" "+str(MSOA_Array[x][1])+" "+str(MSOA_Array[x][2])+" "+str(MSOA_Array[x][3])+" "+str(MSOA_Array[x][4])+" "+str(MSOA_Array[x][5])+"\n")
                x+=1
        MSOA_Output_File.closed
        print("MSOAs are found and exported to file. See "+str(placename)+"_Nodes/MSOAs.txt for details.")
